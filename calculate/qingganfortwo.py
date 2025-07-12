import os
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter
import pandas as pd

# 🔁 控制是否输出额外的情感分析列（主要正面/负面描述词）
# 输出表格列控制
# ENABLE_EXTRA_COLUMNS = True: 输出四列：旅游要素、情感倾向统计、主要正面描述词、主要负面描述词
# ENABLE_EXTRA_COLUMNS = False: 只输出两列：旅游要素、情感倾向统计
ENABLE_EXTRA_COLUMNS = False  # 修改这个开关即可控制输出格式


# ✅ 提取红色字体关键词
def extract_red_keywords(file_path):
    """
    提取指定Excel文件第一个Sheet页中所有红色字体的单元格内容。
    :param file_path: Excel文件路径
    :return: 红色关键词列表
    """
    print(f"🔍 正在从 {file_path} 中提取红色关键词...")

    red_keywords = []

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows()):
            for cell_idx, cell in enumerate(row):
                if cell.font and cell.font.color:
                    color = cell.font.color.rgb
                    # 判断是否是红色 (FF0000 或 FFFF0000)
                    if color in ['FFFF0000', 'FF0000']:
                        if cell.value:
                            value = str(cell.value).strip()
                            red_keywords.append(value)
                            print(f"🔴 提取到红色关键词: '{value}'（位置：第{row_idx + 1}行，第{cell_idx + 1}列）")
        print(f"✅ 成功提取到 {len(red_keywords)} 个红色关键词。")
    except Exception as e:
        print(f"❌ 读取文件失败 {file_path}: {e}")

    return red_keywords


# ✅ 生成统计结果Excel
# ✅ 生成统计结果Excel
def generate_statistics_excel(red_elements, data_file_path, output_file_path):
    """
    根据红色关键词和数据文件生成统计结果表格。
    :param red_elements: 红色关键词列表
    :param data_file_path: 数据源Excel文件路径
    :param output_file_path: 输出文件路径
    """
    print(f"📊 正在从 {data_file_path} 中统计情感数据...")

    try:
        df = pd.read_excel(data_file_path, sheet_name=0)
    except Exception as e:
        print(f"❌ 无法读取数据文件: {e}")
        return

    if df.shape[1] < 3:
        print("❌ 数据文件列数不足，至少需要三列（旅游要素、评论词、情感态度）")
        return

    stats = []
    unique_elements = set(red_elements)
    print(f"🔢 共识别到 {len(unique_elements)} 个唯一旅游要素：{unique_elements}")

    for element in unique_elements:
        print(f"🔎 开始处理关键词: '{element}'")

        # 筛选第二列（评论词）中包含当前关键词的所有行
        filtered = df[df.iloc[:, 1].astype(str).str.contains(element, case=False, na=False)]

        if filtered.empty:
            print(f"⚠️ 未找到关键词 '{element}' 在第二列的相关数据。")
            continue

        pos_count = 0
        neg_count = 0
        pos_comments = []
        neg_comments = []

        # 遍历每一行，记录正面/负面评论，并收集对应评论词
        for _, row in filtered.iterrows():
            comment = str(row.iloc[1]).strip()
            sentiment = str(row.iloc[2]).strip()

            if sentiment == "正面":
                pos_count += 1
                pos_comments.append(comment)
            elif sentiment == "负面":
                neg_count += 1
                neg_comments.append(comment)

        total = len(filtered)  # 所有匹配评论数（包括中性、无等）

        if total == 0:
            print(f"⚠️ 关键词 '{element}' 没有匹配的评论。")
            continue

        positive_percent = (pos_count / total) * 100
        negative_percent = (neg_count / total) * 100

        parts = []
        if positive_percent > 0:
            parts.append(f"正面（{positive_percent:.0f}%）")
        if negative_percent > 0:
            parts.append(f"负面（{negative_percent:.0f}%）")

        sentiment_str = ", ".join(parts)

        top_pos = ", ".join([item for item, _ in Counter(pos_comments).most_common(5)])
        top_neg = ", ".join([item for item, _ in Counter(neg_comments).most_common(5)])

        stat_data = {
            "旅游要素": element,
            "情感倾向统计": sentiment_str,
        }

        if ENABLE_EXTRA_COLUMNS:
            stat_data["主要正面描述词"] = top_pos
            stat_data["主要负面描述词"] = top_neg

        stats.append(stat_data)
        print(f"📊 关键词 '{element}' 统计完成：")
        print(f"   📊 情感分布：{sentiment_str}")
        if ENABLE_EXTRA_COLUMNS:
            print(f"   ➕ 主要正面描述词：{top_pos}")
            print(f"   ➖ 主要负面描述词：{top_neg}")

    if not stats:
        print("❌ 没有有效的统计数据可输出。")
        return

    # 动态设置列名
    if ENABLE_EXTRA_COLUMNS:
        columns_order = ["旅游要素", "情感倾向统计", "主要正面描述词", "主要负面描述词"]
    else:
        columns_order = ["旅游要素", "情感倾向统计"]

    result_df = pd.DataFrame(stats, columns=columns_order)

    # 保存到Excel
    try:
        result_df.to_excel(output_file_path, index=False, engine='openpyxl')
        print(f"✅ 统计结果已成功保存至: {output_file_path}")
    except Exception as e:
        print(f"❌ 保存文件失败: {e}")


# ✅ 主函数入口
def main():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder = os.path.join(current_dir, '..', 'emotionalInputFile')
    output_folder = os.path.join(current_dir, '..', 'outputfile')

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        print(f"📂 请将待处理文件放入输入文件夹: {input_folder}")
        return

    # 查找符合命名规则的文件
    red_file = None
    data_file = None

    for f in os.listdir(input_folder):
        if f.startswith("1_") and f.endswith(".xlsx") and not f.startswith("~$"):
            red_file = f
        elif f.startswith("2_") and f.endswith(".xlsx") and not f.startswith("~$"):
            data_file = f

    if not red_file:
        print("❌ 未找到以 '1_' 开头的红色关键词提取文件，请确保该文件存在且命名正确。")
        return
    if not data_file:
        print("❌ 未找到以 '2_' 开头的待分析文件，请确保该文件存在且命名正确。")
        return

    file1_path = os.path.join(input_folder, red_file)
    file2_path = os.path.join(input_folder, data_file)

    print(f"\n📄 使用文件提取红色关键词: {red_file}")
    print(f"📄 使用文件做数据分析: {data_file}")

    # 提取红色关键词
    red_elements = extract_red_keywords(file1_path)

    # 示例：将以下关键词全部添加到 red_keywords 列表中
    keywords = [
        "荷花",
        "湖水",
        "荷叶",
        "西湖美景三月天嘞春雨如酒柳如烟",
        "山色空蒙雨亦奇",
        "接天莲叶无穷碧",
        "预约",
        "路线",
        "费",
        "门票",
        "酒店",
        "楼外楼",
        "停车",
        "日落"
    ]

    for keyword in keywords:
        red_elements.append(keyword)
    if not red_elements:
        print("❌ 没有提取到任何红色关键词，请检查输入文件格式或颜色设置。")
        return

    # 生成带时间戳的输出文件名
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    output_file = os.path.join(output_folder, f"statistics_output_{timestamp}.xlsx")
    print(f"📁 输出文件路径: {output_file}")

    generate_statistics_excel(red_elements, file2_path, output_file)


if __name__ == '__main__':
    main()
