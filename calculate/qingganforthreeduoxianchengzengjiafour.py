import os
import re
import time
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter
import pandas as pd
from openai import OpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import json
import signal

# 初始化DeepSeek客户端
client = OpenAI(api_key="sk-b56d299a263d4570a59580b1082a262e", base_url="https://api.deepseek.com")

# 全局变量
continue_processing = True
lock = threading.Lock()
final_stats = []
RUN_KEYWORD_COUNT = 3  # 可设置为数字如 5，或 "all"


def analyze_sentiment(comment):
    """
    使用DeepSeek API分析评论的情感倾向，并提取核心关键词。
    :param comment: 待分析的评论文本
    :return: {'sentiment': '正面'/'负面', 'keywords': ['关键词1', '关键词2']} 或 None
    """
    try:
        start_time = time.time()

        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system",
                 "content": "你是一位旅游评论情感分析助手，专门负责判断小红书上的评论是正面还是负面，并提取出最能表达情绪的关键词。"},
                {"role": "user", "content": (
                    "请根据以下标准判断一条来自小红书的评论的情感倾向并提取关键词：\n\n"
                    "1. 如果评论表达了对游览体验的喜爱、赞美或满意，请返回 '正面' 并提取最能代表这种情感的关键词。\n"
                    "2. 如果评论表达了不满、抱怨或批评，请返回 '负面' 并提取最能代表这种情感的关键词。\n"
                    "3. 关键词应简洁明了，例如：免费、方便、准点、人多、排队、贵等。\n"
                    "4. 不要返回整句话，只提取关键词即可。\n"
                    "5. 如果评论提到了多个关键词，请按重要性排序，取最重要的1~2个。\n\n"
                    "请以严格的JSON格式输出结果，格式为：{'sentiment': '正面/负面', 'keywords': ['关键词1', '关键词2']}。\n"
                    "不要添加其他内容，不要使用Markdown格式。\n\n"
                    f"现在请分析以下评论：\n{comment}"
                )}
            ],
            stream=False
        )

        content = response.choices[0].message.content.strip()
        print(f"🧠 DeepSeek原始输出: {content}")

        # ✅ 添加处理逻辑：将单引号替换为双引号，并修复可能的语法问题
        content = content.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
        content = re.sub(r"(?<!\\)'", '"', content)  # 将所有未转义的单引号替换成双引号

        result = json.loads(content)

        sentiment = result.get("sentiment")
        keywords = result.get("keywords", [])

        if not isinstance(keywords, list):
            keywords = []

        elapsed = time.time() - start_time
        print(f"🔍 分析评论 '{comment[:20]}...' 的情感倾向: {sentiment} | 提取关键词: {keywords}（耗时：{elapsed:.2f}s）")

        if sentiment in ["正面", "负面"] and keywords:
            return {"sentiment": sentiment, "keywords": keywords}
        else:
            return None

    except json.JSONDecodeError as je:
        print(f"❌ JSON解析失败: {je}，原始内容为：'{content}'")
        return None
    except Exception as e:
        print(f"❌ 调用DeepSeek API失败: {e}")
        return None


def process_single_element(element, df_comments, sentiment_stats_map):
    print(f"🔎 开始处理关键词: '{element}'")

    # 从sentiment_stats_map获取该关键词的情感统计信息
    stats = sentiment_stats_map.get(element, {'正面': 0, '负面': 0})
    pos_count = stats['正面']
    neg_count = stats['负面']

    total = pos_count + neg_count
    pos_percentage = (pos_count / total * 100) if total > 0 else 0.0
    neg_percentage = (neg_count / total * 100) if total > 0 else 0.0

    # 筛选出包含当前关键词的所有评论
    filtered_comments = df_comments[df_comments.iloc[:, 0].astype(str).str.contains(element, case=False, na=False)]

    pos_keywords = set()
    neg_keywords = set()

    for _, row in filtered_comments.iterrows():
        if not continue_processing:
            break

        comment = str(row.iloc[0]).strip()
        result = analyze_sentiment(comment)

        if result:
            sentiment = result["sentiment"]
            keywords = result["keywords"]

            if sentiment == "正面":
                pos_keywords.update(keywords)
                if len(pos_keywords) >= 5:
                    pos_keywords = set(list(pos_keywords)[:5])
            elif sentiment == "负面":
                neg_keywords.update(keywords)
                if len(neg_keywords) >= 5:
                    neg_keywords = set(list(neg_keywords)[:5])

    top_pos = ", ".join(pos_keywords) if pos_keywords else ""
    top_neg = ", ".join(neg_keywords) if neg_keywords else ""

    sentiment_stat = f"正面({pos_percentage:.1f}%), 负面({neg_percentage:.1f}%)"

    if top_pos or top_neg:
        result = {
            "旅游要素": element,
            "情感倾向统计": sentiment_stat,
            "主要正面描述词": top_pos,
            "主要负面描述词": top_neg
        }
        print(f"📊 关键词 '{element}' 统计完成：")
        print(f"   ➕ 正面描述词：{top_pos}")
        print(f"   ➖ 负面描述词：{top_neg}")
        return result
    else:
        print(f"🗑️ 删除空行：关键词 '{element}' 的主要正/负面描述词均为空。")
        return None


def read_sentiment_stats(file_path):
    sentiment_stats_map = {}
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        for index, row in df.iterrows():
            element = str(row.iloc[0]).strip()  # 第一列：关键词
            sentiment = str(row.iloc[2]).strip().lower()  # 第三列：情感倾向

            if element not in sentiment_stats_map:
                sentiment_stats_map[element] = {'正面': 0, '负面': 0}

            if sentiment == '正面':
                sentiment_stats_map[element]['正面'] += 1
            elif sentiment == '负面':
                sentiment_stats_map[element]['负面'] += 1

        print("📊 情感统计映射构建完成：")
        for k, v in sentiment_stats_map.items():
            print(f"   📌 '{k}': {v}")

    except Exception as e:
        print(f"❌ 无法读取情感统计数据文件: {e}")
    return sentiment_stats_map


def extract_red_keywords(file_path):
    red_keywords = []
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.color:
                    color = cell.font.color.rgb
                    if color in ['FFFF0000', 'FF0000']:
                        if cell.value:
                            value = str(cell.value).strip()
                            red_keywords.append(value)
                            print(f"🔴 提取到红色关键词: '{value}'")
        print(f"✅ 成功提取到 {len(red_keywords)} 个红色关键词。")
    except Exception as e:
        print(f"❌ 读取文件失败 {file_path}: {e}")
    return red_keywords


def signal_handler(sig, frame):
    global continue_processing
    print("\n🛑 用户中断信号捕获 (Ctrl+C)，正在优雅退出...")
    continue_processing = False


def main():
    global continue_processing

    start_time = time.time()
    print(f"🕒 程序启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)

    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder = os.path.join(current_dir, '..', 'emotionalInputFile')
    output_folder = os.path.join(current_dir, '..', 'outputfile')

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    red_file = None
    comments_file = None
    sentiment_file = None

    for f in os.listdir(input_folder):
        if f.startswith("1_") and f.endswith(".xlsx"):
            red_file = f
        elif f.startswith("2_") and f.endswith(".xlsx"):
            sentiment_file = f
        elif f.startswith("3_") and f.endswith(".xlsx"):
            comments_file = f

    if not red_file:
        print("❌ 未找到以 '1_' 开头的红色关键词提取文件。")
        return
    if not comments_file:
        print("❌ 未找到以 '3_' 开头的评论文件。")
        return
    if not sentiment_file:
        print("❌ 未找到以 '2_' 开头的情感统计文件。")
        return

    file1_path = os.path.join(input_folder, red_file)
    file2_path = os.path.join(input_folder, sentiment_file)
    file3_path = os.path.join(input_folder, comments_file)

    print(f"\n📄 使用文件提取红色关键词: {red_file}")
    print(f"📄 使用文件作为情感统计来源: {sentiment_file}")
    print(f"📄 使用文件作为评论来源: {comments_file}")

    red_elements = extract_red_keywords(file1_path)
    keywords = [
        "荷花",
        "湖水",
        "荷叶",
        "西湖美景三月天嘞春雨如酒柳如烟",
        "山色空蒙雨亦奇",
        "接天莲叶无穷碧",
        "预约",
        "路线",
        "费",
        "门票",
        "酒店",
        "楼外楼",
        "停车",
        "日落"
    ]

    for keyword in keywords:
        red_elements.append(keyword)
    print(f"🔍 提取到 {len(red_elements)} 个红色关键词: {red_elements}")
    if not red_elements:
        print("❌ 没有提取到任何红色关键词，请检查输入文件格式或颜色设置。")
        return

    sentiment_stats_map = read_sentiment_stats(file2_path)

    try:
        df_comments = pd.read_excel(file3_path, sheet_name=0)
    except Exception as e:
        print(f"❌ 无法读取评论数据文件: {e}")
        return

    unique_elements = set(red_elements)
    if RUN_KEYWORD_COUNT != "all":
        try:
            run_count = int(RUN_KEYWORD_COUNT)
            unique_elements = list(unique_elements)[:run_count]
        except ValueError:
            print("⚠️ RUN_KEYWORD_COUNT 设置错误，默认跑全量数据。")

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = []
        for element in unique_elements:
            if not continue_processing:
                break
            future = executor.submit(process_single_element, element, df_comments, sentiment_stats_map)
            futures.append(future)

        for future in as_completed(futures):
            if not continue_processing:
                break
            result = future.result()
            if result:
                with lock:
                    final_stats.append(result)

    if final_stats:
        result_df = pd.DataFrame(final_stats, columns=[
            "旅游要素", "情感倾向统计", "主要正面描述词", "主要负面描述词"
        ])

        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        output_file = os.path.join(output_folder, f"statistics_output_{timestamp}.xlsx")
        print(f"📁 输出文件路径: {output_file}")

        try:
            result_df.to_excel(output_file, index=False, engine='openpyxl')
            print(f"✅ 统计结果已成功保存至: {output_file}")
        except Exception as e:
            print(f"❌ 保存文件失败: {e}")
    else:
        print("⚠️ 没有完整的统计数据可输出。")

    total_elapsed = time.time() - start_time
    print(f"🏁 程序结束，总耗时：{total_elapsed:.2f} 秒")


if __name__ == '__main__':
    main()