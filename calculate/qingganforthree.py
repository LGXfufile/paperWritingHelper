import os
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter
import pandas as pd
from openai import OpenAI

# 初始化DeepSeek客户端
client = OpenAI(api_key="sk-b56d299a263d4570a59580b1082a262e", base_url="https://api.deepseek.com")


def analyze_sentiment(comment):
    """
    使用DeepSeek API分析文本的情感倾向。
    :param comment: 待分析的文本
    :return: 返回'正面', '负面' 或 None（对于中性评论）
    """
    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system",
                 "content": "你是一位旅游评论情感分析助手，专门负责判断小红书上的评论是正面、负面还是中性的。"},
                {"role": "user", "content": (
                    "请根据以下标准判断一条来自小红书的评论的情感倾向：\n\n"
                    "1. 如果评论表达了对游览体验的喜爱、赞美或满意，请返回 '正面'。\n"
                    "2. 如果评论表达了不满、抱怨或批评，请返回 '负面'。\n"
                    "3. 如果评论没有明显的情感倾向或者只是陈述事实，请返回 '中性'。\n"
                    "4. 如果评论内容混合了正面和负面的情绪，请尝试提取核心意图，若主要关注点在正面情绪上，请返回 '正面'；反之则返回 '负面'。\n"
                    "5. 特别注意：如果评论提到了一些不便但整体语气积极，或者虽然有些抱怨但是仍然推荐游览，也请考虑为 '正面'。\n\n"
                    "请只输出一个单词：'正面' 或 '负面'，不要添加其他解释或格式。中性评论无需返回。\n\n"
                    f"现在请判断以下评论的情感倾向：\n{comment}"
                )}
            ],
            stream=False
        )
        sentiment = response.choices[0].message.content.strip().lower()
        print(f"🔍 分析评论 '{comment}' 的情感倾向: {sentiment}")

        if sentiment not in ['正面', '负面']:
            return None  # 对于中性评论，我们不处理它们

        return sentiment
    except Exception as e:
        print(f"❌ 调用DeepSeek API失败: {e}")
        return None


def extract_red_keywords(file_path):
    """
    提取指定Excel文件第一个Sheet页中所有红色字体的单元格内容。
    :param file_path: Excel文件路径
    :return: 红色关键词列表
    """
    red_keywords = []

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows()):
            for cell_idx, cell in enumerate(row):
                if cell.font and cell.font.color:
                    color = cell.font.color.rgb
                    if color in ['FFFF0000', 'FF0000']:
                        if cell.value:
                            value = str(cell.value).strip()
                            red_keywords.append(value)
                            print(f"🔴 提取到红色关键词: '{value}'（位置：第{row_idx + 1}行，第{cell_idx + 1}列）")
        print(f"✅ 成功提取到 {len(red_keywords)} 个红色关键词。")
    except Exception as e:
        print(f"❌ 读取文件失败 {file_path}: {e}")

    return red_keywords


def generate_statistics_excel(red_elements, data_file_path, comments_file_path, output_file_path):
    """
    根据红色关键词和数据文件生成统计结果表格。
    :param red_elements: 红色关键词列表
    :param data_file_path: 数据源Excel文件路径
    :param comments_file_path: 评论Excel文件路径
    :param output_file_path: 输出文件路径
    """
    print(f"📊 正在从 {comments_file_path} 中统计情感数据...")

    try:
        df_comments = pd.read_excel(comments_file_path, sheet_name=0)
    except Exception as e:
        print(f"❌ 无法读取评论数据文件: {e}")
        return

    stats = []
    unique_elements = set(red_elements)
    print(f"🔢 共识别到 {len(unique_elements)} 个唯一旅游要素：{unique_elements}")

    # 在generate_statistics_excel函数内...
    for element in unique_elements:
        print(f"🔎 开始处理关键词: '{element}'")

        # 筛选第一列中包含当前关键词的所有评论
        filtered_comments = df_comments[df_comments.iloc[:, 0].astype(str).str.contains(element, case=False, na=False)]

        if filtered_comments.empty:
            print(f"⚠️ 未找到关键词 '{element}' 在第一列的相关评论。")
            continue

        pos_comments = []
        neg_comments = []

        for _, row in filtered_comments.iterrows():
            comment = str(row.iloc[0]).strip()  # 第一列：评论

            sentiment = analyze_sentiment(comment)

            if sentiment == "正面":
                pos_comments.append(comment)
            elif sentiment == "负面":
                neg_comments.append(comment)

        top_pos = ", ".join([item for item in pos_comments[:5]])  # 只保留前5条正面评论
        top_neg = ", ".join([item for item in neg_comments[:5]])  # 只保留前5条负面评论

        if top_pos or top_neg:
            stats.append({
                "旅游要素": element,
                "主要正面描述词": top_pos,
                "主要负面描述词": top_neg
            })
            print(f"📊 关键词 '{element}' 统计完成：")
            print(f"   ➕ 正面描述词：{top_pos}")
            print(f"   ➖ 负面描述词：{top_neg}")
        else:
            print(f"🗑️ 删除空行：关键词 '{element}' 的主要正/负面描述词均为空。")

    if not stats:
        print("❌ 没有有效的统计数据可输出。")
        return

    result_df = pd.DataFrame(stats, columns=[
        "旅游要素", "主要正面描述词", "主要负面描述词"
    ])

    # 保存到Excel
    try:
        result_df.to_excel(output_file_path, index=False, engine='openpyxl')
        print(f"✅ 统计结果已成功保存至: {output_file_path}")
    except Exception as e:
        print(f"❌ 保存文件失败: {e}")


def main():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder = os.path.join(current_dir, '..', 'emotionalInputFile')
    output_folder = os.path.join(current_dir, '..', 'outputfile')

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 查找符合命名规则的文件
    red_file = None
    data_file = None
    comments_file = None

    for f in os.listdir(input_folder):
        if f.startswith("1_") and f.endswith(".xlsx"):
            red_file = f
        elif f.startswith("2_") and f.endswith(".xlsx"):
            data_file = f
        elif f.startswith("3_") and f.endswith(".xlsx"):
            comments_file = f

    if not red_file:
        print("❌ 未找到以 '1_' 开头的红色关键词提取文件。")
        return
    if not comments_file:
        print("❌ 未找到以 '3_' 开头的评论文件。")
        return

    file1_path = os.path.join(input_folder, red_file)
    file3_path = os.path.join(input_folder, comments_file)

    print(f"\n📄 使用文件提取红色关键词: {red_file}")
    print(f"📄 使用文件作为评论来源: {comments_file}")

    # 提取红色关键词
    red_elements = extract_red_keywords(file1_path)
    if not red_elements:
        print("❌ 没有提取到任何红色关键词，请检查输入文件格式或颜色设置。")
        return

    # 生成带时间戳的输出文件名
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    output_file = os.path.join(output_folder, f"statistics_output_{timestamp}.xlsx")
    print(f"📁 输出文件路径: {output_file}")

    generate_statistics_excel(red_elements, None, file3_path, output_file)


if __name__ == '__main__':
    main()
